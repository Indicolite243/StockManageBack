import importlib.util
import json
import os
import sys
import types
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from types import SimpleNamespace

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

try:
    from xtquant import xtdata
except Exception:
    xtdata = None


ENGINE = None


@dataclass
class PerShare:
    type: str = "stock"
    cost: float = 0.0


@dataclass
class PriceSlippage:
    perc: float = 0.0


@dataclass
class Position:
    amount: float = 0.0
    avg_cost: float = 0.0


@dataclass
class Bar:
    open: float
    high: float
    low: float
    close: float
    volume: float
    raw_open: float
    raw_high: float
    raw_low: float
    raw_close: float


class StrategyLogger:
    def info(self, message):
        print(f"[INFO] {message}")

    def warning(self, message):
        print(f"[WARN] {message}")

    def error(self, message):
        print(f"[ERROR] {message}")


class Portfolio:
    def __init__(self, engine):
        self._engine = engine
        self.starting_cash = float(os.environ.get("BACKTEST_INITIAL_CASH", "1000000"))
        self.cash = self.starting_cash
        self.positions = {}

    @property
    def market_value(self):
        total = 0.0
        for code, position in self.positions.items():
            total += position.amount * self._engine.current_mark_prices.get(code, 0.0)
        return float(total)

    @property
    def total_value(self):
        return float(self.cash + self.market_value)


class MindgoBacktestEngine:
    def __init__(self, strategy_path):
        self.strategy_path = Path(strategy_path).resolve()
        self.data_dir = Path(os.environ.get("BACKTEST_DATA_DIR", Path.cwd())).resolve()
        self.start_date = pd.Timestamp(os.environ.get("BACKTEST_START_DATE", "2020-01-01"))
        self.end_date = pd.Timestamp(os.environ.get("BACKTEST_END_DATE", "2025-01-01"))
        self.benchmark_symbol = os.environ.get("BACKTEST_BENCHMARK", "000300.SH")
        self.log = StrategyLogger()
        self.commission_rate = 0.0
        self.slippage_perc = 0.0
        self.volume_limit_ratio = float(os.environ.get("BACKTEST_VOLUME_LIMIT_RATIO", "1.0"))
        self.risk_free_rate = float(os.environ.get("BACKTEST_RISK_FREE_RATE", "0.03"))
        self.current_dt = self.start_date.to_pydatetime()
        self.current_date = self.start_date
        self.current_phase = "init"
        self.current_bars = {}
        self.current_exec_prices = {}
        self.current_mark_prices = {}
        self.current_remaining_volume = {}
        self.records = {}
        self.trade_records = []
        self.data_cache = {}
        self.benchmark_cache = {}
        self.artifacts = {}
        self.context = SimpleNamespace()
        self.context.portfolio = Portfolio(self)
        self.context.run_params = {
            "start_date": self.start_date.strftime("%Y-%m-%d"),
            "end_date": self.end_date.strftime("%Y-%m-%d"),
        }

    def _load_single_symbol(self, code):
        if code in self.data_cache:
            return self.data_cache[code]

        file_path = self.data_dir / f"{code}.xlsx"
        if not file_path.exists():
            raise FileNotFoundError(f"未找到行情文件: {file_path}")

        df = pd.read_excel(file_path, engine="openpyxl")
        df.columns = [str(col).strip() for col in df.columns]
        if "time" not in df.columns:
            raise ValueError(f"{file_path.name} 缺少 time 列")

        df["time"] = pd.to_datetime(df["time"], errors="coerce")
        df = df.dropna(subset=["time"]).sort_values("time").set_index("time")

        for col in ["open", "high", "low", "close", "volume", "adjustment_nv"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        for col in ["open", "high", "low", "close"]:
            df[col] = df.get(col, np.nan).ffill()
        df["volume"] = df.get("volume", 0).fillna(0.0)

        if "adjustment_nv" in df.columns:
            df["adjustment_nv"] = df["adjustment_nv"].ffill()
            valid_close = df["close"].replace(0, np.nan)
            pre_ratio = df["adjustment_nv"] / valid_close
            pre_ratio = pre_ratio.replace([np.inf, -np.inf], np.nan).ffill().fillna(1.0)
            df["open_pre"] = df["open"] * pre_ratio
            df["high_pre"] = df["high"] * pre_ratio
            df["low_pre"] = df["low"] * pre_ratio
            df["close_pre"] = df["adjustment_nv"]
        else:
            df["open_pre"] = df["open"]
            df["high_pre"] = df["high"]
            df["low_pre"] = df["low"]
            df["close_pre"] = df["close"]

        self.data_cache[code] = df
        return df

    def _load_xtdata_symbol(self, code):
        if code in self.benchmark_cache:
            return self.benchmark_cache[code]
        if xtdata is None:
            raise FileNotFoundError(f"未找到行情文件，且 xtdata 不可用: {code}")

        start_time = (self.start_date - pd.Timedelta(days=10)).strftime("%Y%m%d")
        end_time = self.end_date.strftime("%Y%m%d")
        xtdata.download_history_data(code, period="1d", start_time=start_time, end_time=end_time)
        market = xtdata.get_market_data_ex(
            stock_list=[code],
            period="1d",
            start_time=start_time,
            end_time=end_time,
            count=-1,
        )
        df = market.get(code) if isinstance(market, dict) else None
        if df is None or df.empty:
            raise FileNotFoundError(f"无法从 xtdata 获取基准数据: {code}")

        df = df.copy()
        parsed_index = pd.to_datetime(df.index.astype(str), format="%Y%m%d", errors="coerce")
        if not parsed_index.isna().all():
            df.index = parsed_index
            df = df[~df.index.isna()].sort_index()
        elif "time" in df.columns:
            df["time"] = pd.to_datetime(df["time"], unit="ms", errors="coerce")
            df["time"] = df["time"].dt.floor("D")
            df = df.dropna(subset=["time"]).sort_values("time").set_index("time")
        else:
            df.index = pd.to_datetime(df.index, errors="coerce")
            df = df[~df.index.isna()].sort_index()

        for col in ["open", "high", "low", "close", "volume"]:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors="coerce")

        self.benchmark_cache[code] = df
        return df

    def _load_benchmark_symbol(self, code):
        file_path = self.data_dir / f"{code}.xlsx"
        if file_path.exists():
            return self._load_single_symbol(code)
        return self._load_xtdata_symbol(code)

    def get_price(self, code, start_date=None, end_date=None, bar_count=None, fre_step="1d", fields=None, skip_paused=True, fq=None):
        del fre_step, skip_paused
        df = self._load_single_symbol(code)
        price_df = pd.DataFrame(index=df.index)

        use_pre = fq == "pre"
        mapping = {
            "open": "open_pre" if use_pre else "open",
            "high": "high_pre" if use_pre else "high",
            "low": "low_pre" if use_pre else "low",
            "close": "close_pre" if use_pre else "close",
            "volume": "volume",
        }

        requested_fields = fields or ["open", "high", "low", "close", "volume"]
        for field in requested_fields:
            source_col = mapping.get(field)
            if source_col and source_col in df.columns:
                price_df[field] = df[source_col]

        if end_date:
            price_df = price_df.loc[price_df.index <= pd.Timestamp(end_date)]
        if start_date:
            price_df = price_df.loc[price_df.index >= pd.Timestamp(start_date)]
        if bar_count:
            price_df = price_df.tail(int(bar_count))
        return price_df.copy()

    def set_commission(self, commission):
        self.commission_rate = float(getattr(commission, "cost", 0.0))

    def set_slippage(self, slippage):
        self.slippage_perc = float(getattr(slippage, "perc", 0.0))

    def set_benchmark(self, benchmark_symbol):
        if benchmark_symbol:
            self.benchmark_symbol = str(benchmark_symbol)

    def _get_exec_price(self, code, is_buy):
        raw_price = float(self.current_exec_prices.get(code, 0.0))
        if raw_price <= 0:
            raw_price = float(self.current_mark_prices.get(code, 0.0))
        if raw_price <= 0:
            return 0.0
        half_slippage = self.slippage_perc / 2.0
        return raw_price * (1.0 + half_slippage if is_buy else 1.0 - half_slippage)

    def _consume_trade_volume(self, code, requested_qty):
        remaining = float(self.current_remaining_volume.get(code, 0.0))
        actual_qty = min(float(requested_qty), remaining)
        self.current_remaining_volume[code] = max(0.0, remaining - actual_qty)
        return actual_qty

    def _append_trade(self, code, action, quantity, exec_price, fee):
        self.trade_records.append({
            "date": self.current_date.strftime("%Y-%m-%d"),
            "etf": code,
            "action": action,
            "shares": quantity,
            "price": exec_price,
            "amount": quantity * exec_price,
            "commission": fee,
        })

    def order_value(self, code, delta_value):
        if code not in self.current_bars:
            return None

        delta_value = float(delta_value)
        if abs(delta_value) <= 1e-12:
            return None

        is_buy = delta_value > 0
        exec_price = self._get_exec_price(code, is_buy)
        if exec_price <= 0:
            return None

        position = self.context.portfolio.positions.setdefault(code, Position())
        action = "买入" if is_buy else "卖出"

        if is_buy:
            max_notional = self.context.portfolio.cash / (1.0 + self.commission_rate)
            target_notional = min(delta_value, max_notional)
            if target_notional <= 0:
                return None
            requested_qty = target_notional / exec_price
            quantity = self._consume_trade_volume(code, requested_qty)
            if quantity <= 0:
                return None
            trade_value = quantity * exec_price
            fee = trade_value * self.commission_rate
            total_cost = trade_value + fee
            if total_cost > self.context.portfolio.cash:
                quantity = self.context.portfolio.cash / (exec_price * (1.0 + self.commission_rate))
                quantity = self._consume_trade_volume(code, quantity)
                if quantity <= 0:
                    return None
                trade_value = quantity * exec_price
                fee = trade_value * self.commission_rate
                total_cost = trade_value + fee
            previous_cost = position.avg_cost * position.amount
            position.amount += quantity
            position.avg_cost = (previous_cost + trade_value + fee) / position.amount if position.amount > 0 else 0.0
            self.context.portfolio.cash -= total_cost
        else:
            available_amount = max(position.amount, 0.0)
            if available_amount <= 0:
                return None
            requested_qty = min(abs(delta_value) / exec_price, available_amount)
            quantity = self._consume_trade_volume(code, requested_qty)
            quantity = min(quantity, available_amount)
            if quantity <= 0:
                return None
            trade_value = quantity * exec_price
            fee = trade_value * self.commission_rate
            self.context.portfolio.cash += trade_value - fee
            position.amount -= quantity
            if position.amount <= 1e-12:
                self.context.portfolio.positions.pop(code, None)
            else:
                position.avg_cost = position.avg_cost

        self._append_trade(code, action, quantity, exec_price, fee)
        return SimpleNamespace(code=code, action=action, amount=quantity, price=exec_price)

    def order_target(self, code, target_amount):
        position = self.context.portfolio.positions.get(code)
        current_amount = position.amount if position else 0.0
        delta_amount = float(target_amount) - current_amount
        reference_price = float(self.current_mark_prices.get(code, 0.0))
        return self.order_value(code, delta_amount * reference_price)

    def order_target_value(self, code, target_value):
        current_value = 0.0
        if code in self.context.portfolio.positions:
            current_value = self.context.portfolio.positions[code].amount * float(self.current_mark_prices.get(code, 0.0))
        return self.order_value(code, float(target_value) - current_value)

    def order_target_percent(self, code, target_percent):
        target_value = self.context.portfolio.total_value * float(target_percent)
        return self.order_target_value(code, target_value)

    def order_percent(self, code, percent):
        delta_value = self.context.portfolio.total_value * float(percent)
        return self.order_value(code, delta_value)

    def record(self, **kwargs):
        day_key = self.current_date.strftime("%Y-%m-%d")
        self.records.setdefault(day_key, {})
        self.records[day_key].update(kwargs)

    def build_calendar(self, symbols):
        frames = [self._load_single_symbol(code) for code in symbols]
        calendar = frames[0].index
        for frame in frames[1:]:
            calendar = calendar.intersection(frame.index)
        calendar = calendar[(calendar >= self.start_date) & (calendar <= self.end_date)]
        return list(calendar)

    def build_bars(self, symbols, current_date):
        bar_dict = {}
        for code in symbols:
            frame = self._load_single_symbol(code)
            row = frame.loc[current_date]
            bar_dict[code] = Bar(
                open=float(row["open_pre"]),
                high=float(row["high_pre"]),
                low=float(row["low_pre"]),
                close=float(row["close_pre"]),
                volume=float(row["volume"]),
                raw_open=float(row["open"]),
                raw_high=float(row["high"]),
                raw_low=float(row["low"]),
                raw_close=float(row["close"]),
            )
        return bar_dict

    def _prepare_daily_state(self, current_date, symbols):
        self.current_date = pd.Timestamp(current_date)
        self.current_bars = self.build_bars(symbols, self.current_date)
        self.current_exec_prices = {
            code: (bar.close if bar.close > 0 else bar.raw_close)
            for code, bar in self.current_bars.items()
        }
        self.current_mark_prices = {
            code: (bar.close if bar.close > 0 else bar.raw_close)
            for code, bar in self.current_bars.items()
        }
        self.current_remaining_volume = {
            code: max(0.0, bar.volume * self.volume_limit_ratio)
            for code, bar in self.current_bars.items()
        }

    def _set_phase_time(self, phase):
        if phase == "before_trading":
            dt = self.current_date.replace(hour=9, minute=0, second=0).to_pydatetime()
        elif phase == "handle_bar":
            dt = self.current_date.replace(hour=9, minute=31, second=0).to_pydatetime()
        elif phase == "after_trading":
            dt = self.current_date.replace(hour=15, minute=30, second=0).to_pydatetime()
        else:
            dt = self.current_date.to_pydatetime()
        self.current_phase = phase
        self.current_dt = dt
        self.context.current_dt = dt
        self.context.current_phase = phase

    def _get_series_value_on_or_before(self, series, current_date):
        current_ts = pd.Timestamp(current_date)
        eligible = series.loc[series.index <= current_ts]
        if len(eligible) == 0:
            eligible = series.loc[series.index >= current_ts]
        if len(eligible) == 0:
            raise KeyError(f"找不到 {current_ts.strftime('%Y-%m-%d')} 对应的行情数据")
        return float(eligible.iloc[-1] if eligible.index[-1] <= current_ts else eligible.iloc[0])

    def export_results(self):
        if not self.records:
            raise RuntimeError("策略运行结束后没有记录任何净值数据")

        ordered_dates = sorted(self.records.keys())
        start_cash = self.context.portfolio.starting_cash
        benchmark_df = self._load_benchmark_symbol(self.benchmark_symbol)

        first_trade_date = pd.Timestamp(ordered_dates[0])
        benchmark_close_series = pd.to_numeric(benchmark_df["close"], errors="coerce").dropna()
        benchmark_history = benchmark_close_series.loc[benchmark_close_series.index < first_trade_date]
        benchmark_base = float(benchmark_history.iloc[-1]) if len(benchmark_history) > 0 else None

        strategy_returns = []
        benchmark_returns = []
        excess_returns = []

        for date_str in ordered_dates:
            record = self.records[date_str]
            net_value = float(record.get("net_value", start_cash))
            strategy_ret = (net_value / start_cash - 1.0) * 100.0
            strategy_returns.append(round(strategy_ret, 2))

            bench_close = self._get_series_value_on_or_before(benchmark_close_series, date_str)
            if benchmark_base is None:
                benchmark_base = bench_close
            benchmark_ret = (bench_close / benchmark_base - 1.0) * 100.0
            benchmark_returns.append(round(benchmark_ret, 2))
            excess_returns.append(round(strategy_ret - benchmark_ret, 2))

        strategy_nav = np.array([1 + ret / 100.0 for ret in strategy_returns], dtype=float)
        benchmark_nav = np.array([1 + ret / 100.0 for ret in benchmark_returns], dtype=float)
        nv_series = strategy_nav
        peak = np.maximum.accumulate(nv_series)
        drawdown = (nv_series - peak) / peak
        max_drawdown = abs(float(np.min(drawdown))) * 100 if len(drawdown) else 0.0

        if len(strategy_returns) > 1:
            daily_returns = pd.Series(strategy_nav).pct_change().dropna()
            benchmark_daily_returns = pd.Series(benchmark_nav).pct_change().dropna()
            paired = pd.concat([daily_returns, benchmark_daily_returns], axis=1, join="inner")
            paired.columns = ["strategy", "benchmark"]
        else:
            daily_returns = pd.Series(dtype=float)
            benchmark_daily_returns = pd.Series(dtype=float)
            paired = pd.DataFrame(columns=["strategy", "benchmark"])

        n_daily = len(daily_returns)
        if n_daily > 0:
            annual_factor = 250.0 / n_daily
            daily_mean = float(daily_returns.mean())
            volatility = float(np.sqrt(annual_factor * np.square(daily_returns - daily_mean).sum()))
            win_rate = float((daily_returns > 0).mean())
        else:
            volatility = 0.0
            win_rate = 0.0

        if len(paired) > 1 and paired["benchmark"].var() > 0:
            covariance = paired["strategy"].cov(paired["benchmark"])
            beta = covariance / paired["benchmark"].var()
            active_return = paired["strategy"] - paired["benchmark"]
        else:
            beta = 0.0
            active_return = pd.Series(dtype=float)

        total_return = strategy_returns[-1]
        annual_return = ((1 + total_return / 100.0) ** (250.0 / len(strategy_returns)) - 1) * 100 if strategy_returns else 0.0
        benchmark_total_return = benchmark_returns[-1] if benchmark_returns else 0.0
        benchmark_annual_return = ((1 + benchmark_total_return / 100.0) ** (250.0 / len(benchmark_returns)) - 1) * 100 if benchmark_returns else 0.0

        annual_return_decimal = annual_return / 100.0
        benchmark_annual_return_decimal = benchmark_annual_return / 100.0
        risk_free_rate = self.risk_free_rate

        alpha = annual_return_decimal - risk_free_rate - beta * (benchmark_annual_return_decimal - risk_free_rate)
        sharpe = (annual_return_decimal - risk_free_rate) / volatility if volatility > 0 else 0.0

        if len(active_return) > 1:
            active_mean = float(active_return.mean())
            tracking_error = float(np.sqrt((250.0 / (len(active_return) - 1)) * np.square(active_return - active_mean).sum()))
        else:
            tracking_error = 0.0

        information_ratio = ((annual_return_decimal - benchmark_annual_return_decimal) / tracking_error) if tracking_error > 0 else 0.0

        if len(paired) > 0:
            downside_mask = paired["strategy"] < paired["benchmark"]
            downside_diff = (paired["strategy"] - paired["benchmark"])[downside_mask]
            downside_risk = float(np.sqrt((250.0 / len(paired)) * np.square(downside_diff).sum()))
        else:
            downside_risk = 0.0

        sortino = (annual_return_decimal - risk_free_rate) / downside_risk if downside_risk > 0 else 0.0
        total_commission = sum(float(trade.get("commission", 0.0)) for trade in self.trade_records)
        buy_trades = sum(1 for trade in self.trade_records if trade.get("action") == "买入")
        sell_trades = sum(1 for trade in self.trade_records if trade.get("action") == "卖出")
        total_turnover = sum(float(trade.get("amount", 0.0)) for trade in self.trade_records)
        turnover_ratio = (total_turnover / start_cash) * 100 if start_cash else 0.0
        final_net_value = float(self.context.portfolio.total_value)
        metrics = {
            "total_return": f"{total_return:.2f}%",
            "benchmark_return": f"{benchmark_total_return:.2f}%",
            "annual_return": f"{annual_return:.2f}%",
            "benchmark_annual_return": f"{benchmark_annual_return:.2f}%",
            "max_drawdown": f"{max_drawdown:.2f}%",
            "sharpe_ratio": f"{sharpe:.2f}",
            "sortino_ratio": f"{sortino:.2f}",
            "alpha": f"{alpha:.2f}",
            "beta": f"{beta:.2f}",
            "volatility": f"{volatility:.2f}",
            "tracking_error": f"{tracking_error:.2f}",
            "information_ratio": f"{information_ratio:.2f}",
            "downside_risk": f"{downside_risk:.2f}",
            "win_rate": f"{win_rate * 100:.2f}%",
            "final_net_value": f"{final_net_value:,.2f}",
            "trade_count": str(len(self.trade_records)),
            "buy_trade_count": str(buy_trades),
            "sell_trade_count": str(sell_trades),
            "turnover_ratio": f"{turnover_ratio:.2f}%",
            "total_commission": f"{total_commission:,.2f}",
        }

        self._export_markdown()
        self._export_excel()

        json_payload = {
            "dates": ordered_dates,
            "strategy": strategy_returns,
            "benchmark": benchmark_returns,
            "excess": excess_returns,
            "_skip_normalization": True,
            "benchmark_symbol": self.benchmark_symbol,
            "engine": {
                "engine_type": "mindgo_runner",
                "slippage_perc": self.slippage_perc,
                "commission_rate": self.commission_rate,
                "volume_limit_ratio": self.volume_limit_ratio,
                "risk_free_rate": self.risk_free_rate,
            },
            "artifacts": self.artifacts,
            "metrics": metrics,
        }

        with open(self.data_dir / "strategy_performance.json", "w", encoding="utf-8") as f:
            json.dump(json_payload, f, ensure_ascii=False, indent=2)

    def _export_markdown(self):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        report_path = self.data_dir / f"ETF_strategy_report_{timestamp}.md"
        final_value = self.context.portfolio.total_value
        pnl = final_value - self.context.portfolio.starting_cash
        lines = [
            "# ETF策略回测报告",
            f"**生成时间**: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            "",
            "## 1. 绩效摘要",
            f"- **初始资金**: {self.context.portfolio.starting_cash:,.2f}",
            f"- **最终资金**: {final_value:,.2f}",
            f"- **总收益率**: {(pnl / self.context.portfolio.starting_cash) * 100:.2f}%",
            f"- **总盈亏额**: {pnl:,.2f}",
            "",
            "## 2. 交易记录",
            "| 日期 | 标的 | 操作 | 价格 | 数量 | 金额 | 佣金 |",
            "|---|---|---|---|---|---|---|",
        ]
        for trade in self.trade_records:
            lines.append(
                f"| {trade['date']} | {trade['etf']} | {trade['action']} | "
                f"{trade['price']:.4f} | {trade['shares']:.4f} | {trade['amount']:.2f} | {trade['commission']:.2f} |"
            )
        report_path.write_text("\n".join(lines), encoding="utf-8")
        self.artifacts["markdown_report"] = str(report_path)

    def _export_excel(self):
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        excel_path = self.data_dir / f"ETF_strategy_trades_{timestamp}.xlsx"
        wb = Workbook()
        ws = wb.active
        ws.title = "所有交易"
        headers = ["日期", "代码", "操作", "数量", "价格", "金额", "佣金"]
        ws.append(headers)
        fill = PatternFill(start_color="CCE5FF", fill_type="solid")
        font = Font(bold=True)
        for cell in ws[1]:
            cell.fill = fill
            cell.font = font
        for trade in self.trade_records:
            ws.append([
                trade["date"],
                trade["etf"],
                trade["action"],
                trade["shares"],
                trade["price"],
                trade["amount"],
                trade["commission"],
            ])
        wb.save(excel_path)
        self.artifacts["trade_excel"] = str(excel_path)


def get_datetime():
    return ENGINE.current_dt


def get_price(*args, **kwargs):
    return ENGINE.get_price(*args, **kwargs)


def set_commission(*args, **kwargs):
    return ENGINE.set_commission(*args, **kwargs)


def set_slippage(*args, **kwargs):
    return ENGINE.set_slippage(*args, **kwargs)


def set_benchmark(*args, **kwargs):
    return ENGINE.set_benchmark(*args, **kwargs)


def order_value(*args, **kwargs):
    return ENGINE.order_value(*args, **kwargs)


def order_target(*args, **kwargs):
    return ENGINE.order_target(*args, **kwargs)


def order_target_value(*args, **kwargs):
    return ENGINE.order_target_value(*args, **kwargs)


def order_target_percent(*args, **kwargs):
    return ENGINE.order_target_percent(*args, **kwargs)


def order_percent(*args, **kwargs):
    return ENGINE.order_percent(*args, **kwargs)


def record(**kwargs):
    return ENGINE.record(**kwargs)


def install_mindgo_shim():
    shim = types.ModuleType("mindgo_api")
    shim.get_datetime = get_datetime
    shim.get_price = get_price
    shim.set_commission = set_commission
    shim.set_slippage = set_slippage
    shim.set_benchmark = set_benchmark
    shim.order_value = order_value
    shim.order_target = order_target
    shim.order_target_value = order_target_value
    shim.order_target_percent = order_target_percent
    shim.order_percent = order_percent
    shim.record = record
    shim.PerShare = PerShare
    shim.PriceSlippage = PriceSlippage
    shim.log = ENGINE.log
    sys.modules["mindgo_api"] = shim


def load_strategy_module(strategy_path):
    spec = importlib.util.spec_from_file_location("uploaded_mindgo_strategy", strategy_path)
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    spec.loader.exec_module(module)
    return module


def run(strategy_path):
    global ENGINE
    ENGINE = MindgoBacktestEngine(strategy_path)
    install_mindgo_shim()
    strategy_module = load_strategy_module(strategy_path)

    if not hasattr(strategy_module, "init"):
        raise RuntimeError("MindGo 策略缺少 init(context) 入口")

    strategy_module.init(ENGINE.context)
    symbols = list(getattr(ENGINE.context, "valid_etfs", None) or getattr(ENGINE.context, "etf_list", []))
    if not symbols:
        raise RuntimeError("策略初始化后未提供有效标的列表")

    calendar = ENGINE.build_calendar(symbols)
    for current_date in calendar:
        ENGINE._prepare_daily_state(current_date, symbols)

        ENGINE._set_phase_time("before_trading")
        if hasattr(strategy_module, "before_trading"):
            strategy_module.before_trading(ENGINE.context)

        ENGINE._set_phase_time("handle_bar")
        if hasattr(strategy_module, "handle_bar"):
            strategy_module.handle_bar(ENGINE.context, ENGINE.current_bars)

        ENGINE._set_phase_time("after_trading")
        if hasattr(strategy_module, "after_trading"):
            strategy_module.after_trading(ENGINE.context)

        day_key = ENGINE.current_date.strftime("%Y-%m-%d")
        ENGINE.records.setdefault(day_key, {})
        ENGINE.records[day_key].setdefault("net_value", ENGINE.context.portfolio.total_value)

    ENGINE.export_results()


if __name__ == "__main__":
    if len(sys.argv) < 2:
        raise SystemExit("Usage: python mindgo_runner.py <strategy_path>")
    run(sys.argv[1])
